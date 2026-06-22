import urllib.request
import json
import os
import datetime
import sys
import time
import re
import openpyxl

# Configure python path to find modules in the parent directory
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import SessionLocal
from models.user import User
from models.transaction import Transaction
from models.invoice import Invoice

BASE_URL = "http://127.0.0.1:8000"

def make_request(method, endpoint, payload=None, token=None, return_headers=False):
    import urllib.error
    url = f"{BASE_URL}{endpoint}"
    data = None
    headers = {}
    if payload is not None:
        data = json.dumps(payload).encode('utf-8')
        headers['Content-Type'] = 'application/json'
    if token is not None:
        headers['Authorization'] = f"Bearer {token}"
        
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as response:
            status = response.getcode()
            body = response.read()
            resp_headers = response.info()
            if return_headers:
                return status, body, resp_headers
            
            try:
                decoded_body = body.decode('utf-8')
                return status, json.loads(decoded_body)
            except Exception:
                return status, body
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8')
        try:
            parsed_body = json.loads(body)
        except Exception:
            parsed_body = body
        if return_headers:
            return e.code, parsed_body, e.headers
        return e.code, parsed_body
    except Exception as e:
        if return_headers:
            return 500, str(e), {}
        return 500, str(e)

def cleanup_db():
    db = SessionLocal()
    try:
        # Delete test users
        test_emails = ["export_userA@example.com", "export_userB@example.com"]
        users = db.query(User).filter(User.email.in_(test_emails)).all()
        user_ids = [u.id for u in users]
        
        if user_ids:
            # Delete transactions belonging to test users
            tx_count = db.query(Transaction).filter(Transaction.user_id.in_(user_ids)).delete(synchronize_session=False)
            # Delete invoices belonging to test users
            inv_count = db.query(Invoice).filter(Invoice.user_id.in_(user_ids)).delete(synchronize_session=False)
            # Delete test users
            user_count = db.query(User).filter(User.id.in_(user_ids)).delete(synchronize_session=False)
            db.commit()
            print(f"Cleaned up {user_count} test users, {tx_count} transactions, and {inv_count} invoices.")
        else:
            print("No test users to clean up.")
    except Exception as e:
        print(f"Cleanup error: {e}")
        db.rollback()
    finally:
        db.close()

def run_tests():
    print("Starting Phase 10 Report Export Engine verification tests...\n")
    
    # 0. Clean database & exports
    cleanup_db()
    
    # 1. Register & Login User A
    print("Registering and logging in User A...")
    status, res = make_request("POST", "/register", {
        "name": "User A ExportTest",
        "email": "export_userA@example.com",
        "password": "exportpassword1"
    })
    assert status == 200
    user_a_id = res.get("id")
    
    status, login_res = make_request("POST", "/login", {
        "email": "export_userA@example.com",
        "password": "exportpassword1"
    })
    assert status == 200
    token_a = login_res.get("access_token")

    # 2. Empty-State PDF Export Test
    print("Testing Empty-State PDF export for User A...")
    status, pdf_bytes, headers = make_request("GET", "/export-pdf", token=token_a, return_headers=True)
    assert status == 200
    assert len(pdf_bytes) > 0
    assert pdf_bytes.startswith(b"%PDF-")
    
    # Verify Content-Disposition filename format
    cd_header = headers.get("Content-Disposition", "")
    print(f"Content-Disposition: {cd_header}")
    # Filename must match user_a_id and YYYYMMDD_HHMMSS
    pattern = rf'attachment;\s*filename="?user_{user_a_id}_financial_report_\d{{8}}_\d{{6}}\.pdf"?'
    assert re.match(pattern, cd_header, re.IGNORECASE)
    
    # Save PDF locally and check for empty-state text
    temp_pdf_path = "temp_empty_a.pdf"
    with open(temp_pdf_path, "wb") as f:
        f.write(pdf_bytes)
    
    # Read PDF text using pdfplumber to assert empty placeholders exist
    import pdfplumber
    with pdfplumber.open(temp_pdf_path) as pdf:
        pdf_text = ""
        for page in pdf.pages:
            pdf_text += page.extract_text() or ""
            
    print(f"Extracted PDF Text:\n{pdf_text}\n")
    assert "No transactions found." in pdf_text
    assert "No invoices found." in pdf_text
    
    os.remove(temp_pdf_path)
    print("[OK] Empty-State PDF Export passed.")

    # 3. Empty-State Excel Export Test
    print("Testing Empty-State Excel export for User A...")
    status, xlsx_bytes, headers = make_request("GET", "/export-excel", token=token_a, return_headers=True)
    assert status == 200
    assert len(xlsx_bytes) > 0
    
    # Verify Excel Content-Disposition filename format
    cd_header = headers.get("Content-Disposition", "")
    print(f"Content-Disposition: {cd_header}")
    pattern = rf'attachment;\s*filename="?user_{user_a_id}_financial_report_\d{{8}}_\d{{6}}\.xlsx"?'
    assert re.match(pattern, cd_header, re.IGNORECASE)
    
    # Save Excel locally and verify workbook sheets and headers
    temp_xlsx_path = "temp_empty_a.xlsx"
    with open(temp_xlsx_path, "wb") as f:
        f.write(xlsx_bytes)
        
    wb = openpyxl.load_workbook(temp_xlsx_path)
    sheet_names = wb.sheetnames
    print(f"Sheets found: {sheet_names}")
    # Verify workbook sheet count (exactly 7 sheets)
    assert len(sheet_names) == 7
    expected_sheets = ["Transactions", "Invoices", "Analytics", "Forecast", "Health", "Fraud", "Tax"]
    for sheet in expected_sheets:
        assert sheet in sheet_names
        
    # Verify Header rows and Metadata Section are present even when empty
    ws_tx = wb["Transactions"]
    assert ws_tx.cell(row=1, column=1).value == "Transaction ID"
    assert ws_tx.cell(row=1, column=2).value == "Date"
    assert ws_tx.cell(row=1, column=3).value == "Description"
    
    # Verify Metadata Block exists on columns H & I
    assert ws_tx["H1"].value == "REPORT METADATA"
    assert ws_tx["H2"].value == "User Name"
    assert ws_tx["I2"].value == "User A ExportTest"
    assert ws_tx["H3"].value == "User Email"
    assert ws_tx["I3"].value == "export_userA@example.com"
    assert ws_tx["H4"].value == "Generated At"
    assert ws_tx["H5"].value == "Report Type"
    assert ws_tx["H6"].value == "Total Transactions"
    assert ws_tx["I6"].value == 0
    
    # Verify Health Sheet Recommendations
    ws_ht = wb["Health"]
    assert ws_ht.cell(row=1, column=1).value == "Health Metric"
    assert ws_ht.cell(row=6, column=1).value == "Risk Level"
    assert ws_ht.cell(row=7, column=1).value == "Recommendations"
    
    # Verify Fraud Sheet Headers & empty count
    ws_fr = wb["Fraud"]
    assert ws_fr.cell(row=2, column=1).value == "Total Alerts"
    assert ws_fr.cell(row=2, column=2).value == 0
    assert ws_fr.cell(row=5, column=1).value == "Transaction ID"
    
    # Verify Tax Sheet Disclaimer & empty stats
    ws_tax = wb["Tax"]
    assert ws_tax.cell(row=2, column=1).value == "Taxable Income"
    assert ws_tax.cell(row=2, column=2).value == 0.0
    # Disclaimer block
    found_disclaimer = False
    for r in range(1, 20):
        val = ws_tax.cell(row=r, column=2).value
        if val and "estimates only and should not be considered" in str(val):
            found_disclaimer = True
            break
    assert found_disclaimer
    
    wb.close()
    os.remove(temp_xlsx_path)
    print("[OK] Empty-State Excel Workbook structure verified.")

    # 4. BackgroundTask Cleanup Verification
    print("Testing BackgroundTask cleanup for exported files...")
    # Wait for background task to complete deletion of generated files under exports/
    time.sleep(1)
    exports_files = os.listdir("exports")
    # All files matching 'user_{user_a_id}_financial_report_' should have been removed automatically
    user_a_prefix = f"user_{user_a_id}_financial_report_"
    matching_files = [f for f in exports_files if f.startswith(user_a_prefix)]
    print(f"Leftover files matching User A in exports/: {matching_files}")
    assert len(matching_files) == 0
    print("[OK] Temporary file cleanup verified successfully.")

    # 5. Ingest active data and verify report statistics
    print("Injecting transaction and invoice for User A...")
    db = SessionLocal()
    tx1 = Transaction(
        date=datetime.date(2026, 6, 23),
        description="[ExportTest] Consultation Fee",
        amount=1500.00,
        category="Income",
        type="Income",
        user_id=user_a_id
    )
    inv1 = Invoice(
        vendor="[ExportTest] Office Depot",
        invoice_number="INV-E1",
        invoice_date=datetime.date(2026, 6, 22),
        total_amount=100.00,
        gst_number="07Depot1234A1Z1",
        category="Office Expenses",
        file_path="uploads/invoices/depot.pdf",
        ocr_text="Office Depot total: 100.00",
        user_id=user_a_id
    )
    db.add(tx1)
    db.add(inv1)
    db.commit()
    db.close()
    
    # Export Excel and verify counts changed
    status, xlsx_bytes, headers = make_request("GET", "/export-excel", token=token_a, return_headers=True)
    assert status == 200
    temp_xlsx_path = "temp_active_a.xlsx"
    with open(temp_xlsx_path, "wb") as f:
        f.write(xlsx_bytes)
        
    wb = openpyxl.load_workbook(temp_xlsx_path)
    ws_tx = wb["Transactions"]
    # Check that row 2 contains our transaction
    assert ws_tx.cell(row=2, column=3).value == "[ExportTest] Consultation Fee"
    assert ws_tx.cell(row=2, column=4).value == 1500.00
    
    # Metadata count should be 1
    assert ws_tx["I6"].value == 1
    
    # Check Analytics values
    ws_an = wb["Analytics"]
    assert ws_an.cell(row=2, column=2).value == "Total Income"
    assert ws_an.cell(row=2, column=3).value == 1500.00
    
    wb.close()
    os.remove(temp_xlsx_path)
    print("[OK] Active state Excel data counts and values verified.")

    # 6. Multi-User Isolation Boundary Check
    print("Registering and logging in User B...")
    status, res = make_request("POST", "/register", {
        "name": "User B ExportTest",
        "email": "export_userB@example.com",
        "password": "exportpassword2"
    })
    assert status == 200
    user_b_id = res.get("id")
    
    status, login_res_b = make_request("POST", "/login", {
        "email": "export_userB@example.com",
        "password": "exportpassword2"
    })
    assert status == 200
    token_b = login_res_b.get("access_token")
    
    # Inject 1 transaction for User B
    db = SessionLocal()
    tx_b = Transaction(
        date=datetime.date(2026, 6, 23),
        description="[ExportTest] User B Purchase",
        amount=-200.00,
        category="Others",
        type="Expense",
        user_id=user_b_id
    )
    db.add(tx_b)
    db.commit()
    db.close()
    
    # Export Excel for User B
    print("Testing data isolation boundaries in Excel exports...")
    status, xlsx_bytes, headers = make_request("GET", "/export-excel", token=token_b, return_headers=True)
    assert status == 200
    
    # Filename must match User B
    cd_header = headers.get("Content-Disposition", "")
    assert f"user_{user_b_id}_" in cd_header
    
    temp_xlsx_path_b = "temp_active_b.xlsx"
    with open(temp_xlsx_path_b, "wb") as f:
        f.write(xlsx_bytes)
        
    wb = openpyxl.load_workbook(temp_xlsx_path_b)
    ws_tx = wb["Transactions"]
    
    # Check that it contains User B's transaction only
    # Row 1 is header, Row 2 is transaction, Row 3 should be empty
    assert ws_tx.cell(row=2, column=3).value == "[ExportTest] User B Purchase"
    assert ws_tx.cell(row=3, column=3).value is None
    
    # Metadata count should be exactly 1
    assert ws_tx["I6"].value == 1
    
    wb.close()
    os.remove(temp_xlsx_path_b)
    print("[OK] Multi-user data isolation verified successfully (User B report does not leak User A's data).")

    # 7. Clean up database & exports
    print("Running self-cleaning database logic...")
    cleanup_db()
    
    # Final sleep to let background deletion complete
    time.sleep(1)
    print("\nALL PHASE 10 REPORT EXPORT ENGINE TESTS COMPLETED SUCCESSFULLY!")

if __name__ == "__main__":
    run_tests()
