import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import analytics modules
from analytics.tax_estimator import estimate_tax, generate_tax_insights
from analytics.tax_summary import get_gst_summary
from analytics.health_score import calculate_health_score, get_financial_grade
from analytics.risk_analysis import analyze_risk
from analytics.recommendations import generate_recommendations
from analytics.forecast import forecast_summary, calculate_cash_runway, build_monthly_series
from analytics.cashflow import calculate_cashflow
from analytics.profit_loss import profit_loss
from analytics.category_analysis import category_summary
from reports.pdf_report import get_merged_alerts

def format_recommendation(r):
    if isinstance(r, dict):
        return r.get("message", str(r))
    return str(r)


def add_excel_metadata(ws, user, transactions, invoices):
    ws["H1"] = "REPORT METADATA"
    ws["H1"].font = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
    ws["H2"] = "User Name"
    ws["I2"] = user.name or "N/A"
    ws["H3"] = "User Email"
    ws["I3"] = user.email
    ws["H4"] = "Generated At"
    ws["I4"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["H5"] = "Report Type"
    ws["I5"] = "Excel Financial Export"
    ws["H6"] = "Total Transactions"
    ws["I6"] = len(transactions)
    ws["H7"] = "Total Invoices"
    ws["I7"] = len(invoices)
    
    # Style metadata columns
    for row in range(1, 8):
        ws.cell(row=row, column=8).font = Font(name="Calibri", size=9, bold=True, color="475569")
        ws.cell(row=row, column=9).font = Font(name="Calibri", size=9, color="1E293B")

def apply_header_style(ws, max_col):
    blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color="CBD5E1"),
        right=Side(style='thin', color="CBD5E1"),
        top=Side(style='thin', color="CBD5E1"),
        bottom=Side(style='thin', color="CBD5E1")
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 24

def autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # We only fit columns up to I (9)
        if col[0].column > 9:
            continue
            
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

def generate_excel_report(user, transactions, invoices):
    os.makedirs("exports", exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"user_{user.id}_financial_report_{timestamp}.xlsx"
    file_path = os.path.join("exports", filename)

    wb = openpyxl.Workbook()

    # --- Fetch Analytics Data ---
    tax_est = estimate_tax(transactions)
    tax_ins = generate_tax_insights(transactions)
    gst_sum = get_gst_summary(invoices)
    
    health_data = calculate_health_score(transactions)
    health_score = health_data["health_score"]
    grade_info = get_financial_grade(health_score)
    risk_info = analyze_risk(transactions)
    recs = generate_recommendations(transactions)
    
    runway_info = calculate_cash_runway(transactions)
    forecast_data = forecast_summary(transactions)
    
    cashflow_data = calculate_cashflow(transactions)
    pnl_data = profit_loss(transactions)
    cat_data = category_summary(transactions)
    monthly_series = build_monthly_series(transactions)
    
    fraud_alerts = get_merged_alerts(transactions)
    total_alerts = len(fraud_alerts)
    avg_fraud_score = round(sum(a["fraud_score"] for a in fraud_alerts) / total_alerts, 2) if total_alerts > 0 else 0.0

    # Define common format styles
    currency_font = Font(name="Calibri", size=11)
    currency_align = Alignment(horizontal="right")
    
    # 1. Transactions Sheet
    ws_tx = wb.active
    ws_tx.title = "Transactions"
    ws_tx.append(["Transaction ID", "Date", "Description", "Amount", "Category", "Type"])
    for t in transactions:
        ws_tx.append([
            t.id,
            t.date.strftime("%Y-%m-%d") if t.date else None,
            t.description,
            t.amount,
            t.category,
            t.type
        ])
    apply_header_style(ws_tx, 6)
    add_excel_metadata(ws_tx, user, transactions, invoices)
    autofit_columns(ws_tx)

    # 2. Invoices Sheet
    ws_inv = wb.create_sheet(title="Invoices")
    ws_inv.append(["Invoice ID", "Vendor", "Invoice Number", "Invoice Date", "Total Amount", "GST Number", "Category", "File Path", "Created At"])
    for inv in invoices:
        ws_inv.append([
            inv.id,
            inv.vendor,
            inv.invoice_number,
            inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else None,
            inv.total_amount,
            inv.gst_number,
            inv.category,
            inv.file_path,
            inv.created_at.strftime("%Y-%m-%d %H:%M:%S") if inv.created_at else None
        ])
    apply_header_style(ws_inv, 9)
    add_excel_metadata(ws_inv, user, transactions, invoices)
    autofit_columns(ws_inv)

    # 3. Analytics Sheet
    ws_an = wb.create_sheet(title="Analytics")
    ws_an.append(["Section", "Metric / Detail", "Value"])
    
    # P&L Rows
    ws_an.append(["Profit & Loss", "Total Income", tax_est["income"]])
    ws_an.append(["Profit & Loss", "Total Expenses", tax_est["expenses"]])
    ws_an.append(["Profit & Loss", "Net Cashflow", cashflow_data.get("net_cashflow", 0.0)])
    ws_an.append(["Profit & Loss", "Burn Rate", cashflow_data.get("burn_rate", 0.0)])
    
    # Category Analysis Rows
    for cat, amt in cat_data.items():
        ws_an.append(["Category Analysis", cat, amt])
        
    # Monthly Summary Rows
    for month_data in monthly_series:
        m = month_data["month"]
        net_cf = round(month_data["income"] - month_data["expenses"], 2)
        ws_an.append(["Monthly Summary", f"{m} Net Cashflow", net_cf])
        
    apply_header_style(ws_an, 3)
    add_excel_metadata(ws_an, user, transactions, invoices)
    autofit_columns(ws_an)

    # 4. Forecast Sheet
    ws_fc = wb.create_sheet(title="Forecast")
    ws_fc.append(["Forecast Metric", "Value"])
    ws_fc.append(["Predicted Income", forecast_data["predicted_income"]])
    ws_fc.append(["Predicted Expenses", forecast_data["predicted_expenses"]])
    ws_fc.append(["Income Trend", forecast_data["income_trend"]])
    ws_fc.append(["Expense Trend", forecast_data["expenses_trend"]])
    ws_fc.append(["Cash Runway", forecast_data["months_remaining"]])
    ws_fc.append(["Confidence", forecast_data["confidence"].upper()])
    
    apply_header_style(ws_fc, 2)
    add_excel_metadata(ws_fc, user, transactions, invoices)
    autofit_columns(ws_fc)

    # 5. Health Sheet
    ws_ht = wb.create_sheet(title="Health")
    ws_ht.append(["Health Metric", "Value"])
    ws_ht.append(["Health Score", health_score])
    ws_ht.append(["Grade", grade_info["grade"]])
    ws_ht.append(["Savings Ratio", health_data["components"]["raw"]["savings_ratio"]])
    ws_ht.append(["Expense Ratio", health_data["components"]["raw"]["expense_ratio"]])
    ws_ht.append(["Risk Level", risk_info.get("risk_level", "N/A").upper()])
    ws_ht.append(["Recommendations", "; ".join(format_recommendation(r) for r in recs)])
    
    apply_header_style(ws_ht, 2)
    add_excel_metadata(ws_ht, user, transactions, invoices)
    autofit_columns(ws_ht)

    # 6. Fraud Sheet
    ws_fr = wb.create_sheet(title="Fraud")
    ws_fr.append(["Fraud Metric / Alert Fields", "Value / Detail", "", "", ""])
    ws_fr.append(["Total Alerts", total_alerts])
    ws_fr.append(["Average Fraud Score", avg_fraud_score])
    ws_fr.append([]) # blank separator
    
    # Table headers for Merged Alerts
    ws_fr.append(["Transaction ID", "Severity", "Fraud Score", "Flagged Sources", "Reasons"])
    for alert in fraud_alerts:
        ws_fr.append([
            alert["transaction_id"],
            alert["severity"].upper(),
            alert["fraud_score"],
            ", ".join(alert["sources"]),
            "; ".join(alert["reasons"])
        ])
        
    apply_header_style(ws_fr, 5)
    add_excel_metadata(ws_fr, user, transactions, invoices)
    autofit_columns(ws_fr)

    # 7. Tax Sheet
    ws_tx_sh = wb.create_sheet(title="Tax")
    ws_tx_sh.append(["Tax Metric / Insight Type", "Value / Detail"])
    ws_tx_sh.append(["Taxable Income", tax_est["taxable_income"]])
    ws_tx_sh.append(["Expense Ratio", tax_est["expense_ratio"]])
    ws_tx_sh.append(["Total GST Paid", gst_sum["total_gst_paid"]])
    ws_tx_sh.append(["Explicit GST Invoices", gst_sum["explicit_gst_invoices"]])
    ws_tx_sh.append(["Estimated GST Invoices", gst_sum["estimated_gst_invoices"]])
    ws_tx_sh.append(["Average GST per Invoice", gst_sum["average_gst"]])
    ws_tx_sh.append([]) # blank separator
    
    # Tax Insights
    for insight in tax_ins.get("insights", []):
        ws_tx_sh.append(["Tax Insight", insight])
    ws_tx_sh.append([])
    ws_tx_sh.append(["Disclaimer", "Tax calculations are estimates only and should not be considered professional or legal advice."])
    
    apply_header_style(ws_tx_sh, 2)
    add_excel_metadata(ws_tx_sh, user, transactions, invoices)
    autofit_columns(ws_tx_sh)

    # Save to disk
    wb.save(file_path)
    return file_path
